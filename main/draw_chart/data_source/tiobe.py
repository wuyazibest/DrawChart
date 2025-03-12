#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# @File   : tiobe.py
# @Time   : 2021/8/12 10:48
# @Author : wuyazibest
# @Email  : wuyazibest@163.com
# @Desc   :

import datetime
import logging
import os

import requests
import re
import numpy as np
import pandas as pd

# from lxml import etree
from openpyxl import Workbook, load_workbook

logger = logging.getLogger(__name__)


def default_path():
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), "language_data.xlsx")


def transform_local_time(utc_time_str):
    try:
        split_char = ", "
        utc_time_str = utc_time_str.split(split_char)
        utc_time_str[1] = str(int(utc_time_str[1]) + 1)
        utc_time_str = split_char.join(utc_time_str)
        tt = datetime.datetime.strptime(utc_time_str, "%Y, %m, %d")
        return datetime.datetime.strftime(tt + datetime.timedelta(hours=8), "%Y-%m-%d")
    except Exception as e:
        logger.error(f"utc_time_str:{utc_time_str}  error:{e}")
        return utc_time_str


def request_tiobe():
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/133.0.0.0 Safari/537.36"
        }
    url = 'https://www.tiobe.com/tiobe-index/'
    resp = requests.get(url, headers=headers).text
    
    # html = etree.HTML(resp)
    # data = html.xpath("//article/script[2]")
    # 正则匹配提取数据
    match = re.findall(r"\{name : '(.*?)',data : (.*?)\}", resp)
    data = []
    for i in match:
        data.extend([[i[0], transform_local_time(x[0]), x[1]] for x in re.findall(r'\[Date.UTC\((.*?)\), (.*?)\]', i[1], re.S)])
    return data


def save(data, path=""):
    path = path or default_path()
    exl = Workbook()  # 创建工作簿对象
    sheet = exl.active  # 获取活动的工作表
    # 时间   编程语言   热度
    sheet.append(['programing', 'date', 'data_per'])
    for i in data:
        sheet.append(i)
    exl.save(path)


def reader(path=""):
    path = path or default_path()
    exl = load_workbook(filename=path, data_only=True)
    sheet = exl.active
    return tuple(sheet.values)[1:]


def get_format_data():
    pd.set_option('display.float_format', lambda x: str(x))
    data = reader()
    df = pd.DataFrame(data, columns=['programing', 'date', 'data_per'])
    df = df.pivot('date', 'programing', 'data_per')
    df = df.astype("float")
    df = df.round(2)
    df = df.replace({np.nan: None})
    # df.to_dict()
    return df


if __name__ == '__main__':
    # data = request_tiobe()
    # save(data)
    reader()
    ret = get_format_data()
    print(ret)
